from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import Font

#Replace sample file paths before running script
course_outlines = [
    r'C:\Users\Name\Downloads\Fall 2025_ Course Name 1.html', 
    r'C:\Users\Name\Downloads\Fall 2025_ Course Name 2.html',
    r'C:\Users\Name\Downloads\Fall 2025_ Course Name 3.html'
]

def parse_course_html(html_file_path):
    with open(html_file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'lxml')

    course_code = soup.find('span', class_='outline-courses').get_text(strip=True)
    course_name = soup.find('span', class_='outline-title-full').get_text(strip=True)
    course_description = soup.find('p', class_='cd-content').get_text(strip=True)

    #learning outcomes table always comes first
    learning_outcomes = soup.find('table', class_='multitable').get_text(separator='\n', strip=True)

    #find assessements and activities table
    heading = soup.find('h2', id='assessments_amp_activities')
    grading_table = heading.find_next('table', class_='multitable')

    all_grading_headers = [th.get_text(strip=True) for th in grading_table.find_all("th")]
    grading_headers = [all_grading_headers[0], all_grading_headers[3]]  # name and weight only

    grading_rows = []
    for tr in grading_table.find_all("tr")[1:]:
        all_cells = [td.get_text(strip=True) for td in tr.find_all("td")]
        if all_cells:
            # Get only assessment name and weight (indices 0 and 3)
            cells = [all_cells[0], all_cells[3]]
            grading_rows.append(cells)
        
    df = pd.DataFrame(grading_rows, columns=grading_headers)

    course_dict = {"course_code" : course_code, "course_name": course_name, 
                   "course_description": course_description, 
                   "learning outcomes": learning_outcomes, 
                   "assessments_df": df}
    
    return course_dict


# Collect all course data
all_courses = []
for course_file in course_outlines:
    try:
        course_data = parse_course_html(course_file)
        all_courses.append(course_data)
    except Exception as e:
        print(f"Error processing {course_file}: {e}")
        continue

# Export to Excel with one sheet per course
output_filename = "All_Courses_Outline.xlsx"

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    for course in all_courses:
        course_code = course['course_code']
        course_name = course['course_name']
        course_description = course['course_description']
        learning_outcomes = course['learning outcomes']
        df = course['assessments_df']
        
        # Create a single sheet for this course with course code as sheet name
        # Create an empty sheet first
        pd.DataFrame().to_excel(writer, sheet_name=course_code, index=False)
        
        # Get the workbook and sheet to write data
        workbook = writer.book
        sheet = workbook[course_code]
        
        # Write course info at the top (without header row, with bold labels)
        course_code_cell = sheet.cell(row=1, column=1, value="Course Code")
        course_code_cell.font = Font(bold=True)
        sheet.cell(row=1, column=2, value=course_code)
        
        course_name_cell = sheet.cell(row=2, column=1, value="Course Name")
        course_name_cell.font = Font(bold=True)
        sheet.cell(row=2, column=2, value=course_name)
        
        current_row = 4
        
        # Write assessments dataframe using pandas (includes headers)
        df.to_excel(writer, sheet_name=course_code, index=False, startrow=current_row - 1, startcol=0)
        
        # Update current_row for next section 
        current_row += len(df) + 2
        
        # Course Description
        course_desc_cell = sheet.cell(row=current_row, column=1, value="Course Description")
        course_desc_cell.font = Font(bold=True)
        current_row += 1
        sheet.cell(row=current_row, column=1, value=course_description)
        current_row += 2
        
        #Learning outcomes header
        learning_outcomes_cell = sheet.cell(row=current_row, column=1, value="Learning Outcomes")
        learning_outcomes_cell.font = Font(bold=True)
        current_row += 1
        
        # Learning outcomes
        outcomes_list = [outcome.strip() for outcome in learning_outcomes.split('\n') if outcome.strip()]
        for idx, outcome in enumerate(outcomes_list):
            sheet.cell(row=current_row + idx, column=1, value=outcome)

print(f"Data exported successfully to: {output_filename}")
